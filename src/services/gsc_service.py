import os
import json
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from google.oauth2 import service_account
from googleapiclient.discovery import build
from src.config import config

logger = logging.getLogger(__name__)

class GSCService:
    def __init__(self, service_account_path: Optional[str] = None):
        self.service_account_path = service_account_path or os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
        self.creds = None
        self.gsc_client = None
        self.indexing_client = None
        
        if self.service_account_path and os.path.exists(self.service_account_path):
            try:
                self.creds = service_account.Credentials.from_service_account_file(
                    self.service_account_path,
                    scopes=[
                        'https://www.googleapis.com/auth/webmasters.readonly',
                        'https://www.googleapis.com/auth/indexing'
                    ]
                )
                self.gsc_client = build('searchconsole', 'v1', credentials=self.creds)
                self.indexing_client = build('indexing', 'v3', credentials=self.creds)
            except Exception as e:
                logger.error(f"Failed to initialize GSC clients: {e}")

    def is_available(self) -> bool:
        return self.creds is not None

    def inspect_url(self, site_url: str, inspection_url: str) -> Dict[str, Any]:
        """Inspect a single URL via GSC URL Inspection API."""
        if not self.gsc_client:
            return {"error": "GSC client not initialized"}
        
        try:
            request = {
                'inspectionUrl': inspection_url,
                'siteUrl': site_url,
                'languageCode': 'en-US'
            }
            result = self.gsc_client.urlInspection().index().inspect(body=request).execute()
            return result.get('inspectionResult', {})
        except Exception as e:
            logger.error(f"Error inspecting URL {inspection_url}: {e}")
            return {"error": str(e)}

    def get_search_analytics(self, site_url: str, url: str, days: int = 30) -> Dict[str, Any]:
        """Fetch analytics for a specific URL."""
        if not self.gsc_client:
            return {}
        
        try:
            request = {
                'startDate': (datetime.now().replace(day=1)).strftime('%Y-%m-%d'), # Simplified for demo
                'endDate': datetime.now().strftime('%Y-%m-%d'),
                'dimensions': ['page'],
                'dimensionFilterGroups': [{
                    'filters': [{
                        'dimension': 'page',
                        'operator': 'equals',
                        'expression': url
                    }]
                }]
            }
            # Adjust start date to 'days' ago
            from datetime import timedelta
            start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            request['startDate'] = start_date

            response = self.gsc_client.searchanalytics().query(siteUrl=site_url, body=request).execute()
            rows = response.get('rows', [])
            if rows:
                return rows[0]
            return {"clicks": 0, "impressions": 0, "ctr": 0, "position": 0}
        except Exception as e:
            logger.error(f"Error fetching analytics for {url}: {e}")
            return {}

    def submit_for_indexing(self, url: str) -> bool:
        """Submit URL to Google Indexing API."""
        if not self.indexing_client:
            return False
        
        try:
            body = {
                'url': url,
                'type': 'URL_UPDATED'
            }
            self.indexing_client.urlNotifications().publish(body=body).execute()
            return True
        except Exception as e:
            logger.error(f"Error submitting {url} for indexing: {e}")
            return False

    def generate_excel_report(self, indexed_data: List[Dict], unindexed_data: List[Dict], filename: str):
        """Generate color-coded Excel report with 2 sheets."""
        wb = Workbook()
        
        # Helper for styling
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        indexed_fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")
        unindexed_fill = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")

        def write_sheet(sheet, data, title, color_fill):
            sheet.title = title
            headers = ["URL", "Status", "Reason", "Clicks", "Impressions", "CTR", "Position"]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, item in enumerate(data, 2):
                sheet.cell(row=row_idx, column=1).value = item.get("url")
                sheet.cell(row=row_idx, column=2).value = item.get("status", "Unknown")
                sheet.cell(row=row_idx, column=3).value = item.get("reason", "-")
                sheet.cell(row=row_idx, column=4).value = item.get("clicks", 0)
                sheet.cell(row=row_idx, column=5).value = item.get("impressions", 0)
                sheet.cell(row=row_idx, column=6).value = f"{item.get('ctr', 0):.2%}"
                sheet.cell(row=row_idx, column=7).value = round(item.get("position", 0), 1)
                
                for col in range(1, 8):
                    sheet.cell(row=row_idx, column=col).fill = color_fill

        # Sheet 1: Indexed
        write_sheet(wb.active, indexed_data, "Indexed Pages", indexed_fill)
        
        # Sheet 2: Unindexed
        wb.create_sheet("Unindexed Pages")
        write_sheet(wb["Unindexed Pages"], unindexed_data, "Unindexed Pages", unindexed_fill)
        
        wb.save(filename)
        return filename

    def analyze_sitemap_gaps(self, sitemap_urls: List[str], crawled_urls: List[str]) -> Dict[str, List[str]]:
        """Identify missing URLs in sitemap or orphaned pages on site."""
        sitemap_set = set(sitemap_urls)
        crawled_set = set(crawled_urls)
        
        missing_in_sitemap = list(crawled_set - sitemap_set)
        orphaned_in_sitemap = list(sitemap_set - crawled_set) # Present in sitemap but not found by crawler
        
        return {
            "missing_in_sitemap": missing_in_sitemap,
            "orphaned_in_sitemap": orphaned_in_sitemap
        }
